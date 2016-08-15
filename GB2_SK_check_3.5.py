# Create SK burn checklists before load SK burn
# This script using Orbital Event file, Satellite Visibility Schedule file, and GB2_MANON file
# to check the orbit events and ground contact in order to create SK burn checklist.

# N:\IOdata\FDS\M*** --------------- GB2_MANON file
# N:\IOdata\FDS\orbital_events ---- Orbital Event file
# N:\IOdata\FDS\sts -------------- Satellite Visibility Schedule file
# N:\sateng\Bus Subsystems\Subsys_AOCS\SK Maneuvers ------ SK burn Checklist Templit
# N:\sateng\Bus Subsystems\Subsys_AOCS\SK Maneuvers\M*** ------ saved SK burn Checklist

from datetime import *
import os
from xml.dom import minidom
from tkinter import  messagebox  as  msg
import tkinter as tk
import openpyxl
import heapq

# Define a function to find the SK burn times
# and determine whether there are a least 15 mins between two burns
def findSKburnTimes(MANON_file, t_format):
      xmldoc = minidom.parse(MANON_file)
      variablefile = xmldoc.getElementsByTagName('VariableFile')[0]
      variables = variablefile.getElementsByTagName('Variable')
      burn_times = []
      for var in variables:
            if var.attributes['Type'].value =='Date':
                  burn_times.append(datetime.strptime(var.attributes['Value'].value, t_format))
      if len(burn_times) > 2:
            i = 1
            while i in range(1, len(burn_times)):
                  if burn_times[i] - burn_times[i+1] <= timedelta(minutes = 15):
                        tk.Tk().withdraw()
                        msg.showinfo('message', \
                            'There are two burns which are less than 15 minutes apart from each other')
                        burn_times = None
      return burn_times

#####################################################################################################

def findburnInfos(MANON_file):
      MANON_doc = minidom.parse(MANON_file)
      variablefile = MANON_doc.getElementsByTagName('VariableFile')[0]
      variables = variablefile.getElementsByTagName('Variable')
      thrusters = []
      bias_RPY = []
      for var in variables:
            if var.attributes['Description'].value =='Delta-V':
                  delta_V = var.attributes['Value'].value
            elif '_bias' in var.attributes['Name'].value:
                  bias_RPY.append(var.attributes['Value'].value)
            elif 'Number of cycles for boost' in var.attributes['Description'].value and \
                 var.attributes['Value'].value != '0':
                  thrusters.append(var.attributes['Name'].value[10])

      return [thrusters, bias_RPY, delta_V]       
#################################################################################################


def findOrbitEvent(flightModel, burn_times, t_format, NBK_flag):
      # Orbital Events
      #path_orbit =  'N:/IOdata/FDS/orbital_events'
      path_orbit =  'C:/Python35/SK_project'
      list_orbit = []
      for i in os.listdir(path_orbit):
            if os.path.isfile(os.path.join(path_orbit,i)) and 'Orbital_Event_' in i:
                  list_orbit.append(i)
            
      list_orbit_date = []
      for j in range(0, len(list_orbit)):
            list_orbit_date.append(datetime.fromtimestamp(os.path.getmtime(path_orbit + '/' + list_orbit[j])))

      index_latest_orbit = list_orbit_date.index(max(list_orbit_date))
      sel_orbit = path_orbit + '/' + list_orbit[index_latest_orbit]
      
      orb_doc = minidom.parse(sel_orbit)
      events = orb_doc.getElementsByTagName('events')[0]
      event = events.getElementsByTagName('event')
      six_am = []
      LCZ_entry = []
      LCZ_exit = []
      eclipse_entry = []
      eclipse_exit = []
      noon = []
      sun_intru_entry = []
      sun_intru_exit = []
      if NBK_flag == 0:
          for eve in event:
              if eve.attributes['resourceName'].value == flightModel:
                  if '6am' in eve.attributes['description'].value:
                      six_am.append(datetime.strptime(eve.attributes['start'].value, t_format))
                  elif 'noon' in eve.attributes['description'].value:
                      noon.append(datetime.strptime(eve.attributes['start'].value, t_format))
                  elif 'Earth Penumbra' in eve.attributes['description'].value:
                      eclipse_entry.append(datetime.strptime(eve.attributes['start'].value, t_format))
                      eclipse_exit.append(datetime.strptime(eve.attributes['end'].value, t_format))
                  elif 'Sun for Sun' in eve.attributes['description'].value and 'ESA' in eve.attributes['description'].value:
                      sun_intru_entry.append(datetime.strptime(eve.attributes['start'].value, t_format))
                      sun_intru_exit.append(datetime.strptime(eve.attributes['end'].value, t_format))
          return [sel_orbit, six_am, eclipse_entry, eclipse_exit, noon, sun_intru_entry, sun_intru_exit]
      else:
          for eve in event:
              if eve.attributes['resourceName'].value == flightModel:
                  if 'FDS Low Commandability Zone' in eve.attributes['description'].value:
                      LCZ_entry.append(datetime.strptime(eve.attributes['start'].value, t_format))
                      LCZ_exit.append(datetime.strptime(eve.attributes['end'].value, t_format))
                  elif 'noon' in eve.attributes['description'].value:
                      noon.append(datetime.strptime(eve.attributes['start'].value, t_format))
                  elif 'Earth Penumbra' in eve.attributes['description'].value:
                      eclipse_entry.append(datetime.strptime(eve.attributes['start'].value, t_format))
                      eclipse_exit.append(datetime.strptime(eve.attributes['end'].value, t_format))
                  elif 'Sun for Sun' in eve.attributes['description'].value and 'ESA' in eve.attributes['description'].value:
                      sun_intru_entry.append(datetime.strptime(eve.attributes['start'].value, t_format))
                      sun_intru_exit.append(datetime.strptime(eve.attributes['end'].value, t_format))
          return [LCZ_entry, LCZ_exit, eclipse_entry, eclipse_exit, noon, sun_intru_entry, sun_intru_exit]

      

#########################################################################################################

def findGroundContact(flightModel, burn_times, six_am_bef, six_am_aft, t_format, NBK_flag):
      # Satellite Visibility Schedule
      #path_sched =  'N:/IOdata/FDS/sts'
      path_sched =  'C:/Python35/SK_project'
      list_sched = []
      for i in os.listdir(path_sched):
            if os.path.isfile(os.path.join(path_sched,i)) and 'Satellite_Visibility_Schedule' in i:
                  list_sched.append(i)
            
      list_sched_date = []
      for j in range(0, len(list_sched)):
            list_sched_date.append(datetime.fromtimestamp(os.path.getmtime(path_sched + '/' + list_sched[j])))

      index_latest_sched = list_sched_date.index(max(list_sched_date))
      sel_sched = path_sched + '/' + list_sched[index_latest_sched]
      
      sched_doc = minidom.parse(sel_sched)
      sts = sched_doc.getElementsByTagName('sts')[0]
      contacts = sts.getElementsByTagName('contact')
      six_am_bef_contacts = []         # list of path of 6 am to disable if it is in view
      six_am_aft_contacts = []         # list of path of 6 am to re-enable if it is in view
      burn_contacts = []               # list of paths of burns if the burns are in view
      contacts_aft_burn = []           # list of paths after SK burn(s)
      cleanup_contacts = []            # list of path for clean up
      for contact in contacts:
            if contact.attributes['satName'].value == flightModel:
                  if datetime.strptime(contact.attributes['start'].value, t_format) <= six_am_bef and \
                      datetime.strptime(contact.attributes['end'].value, t_format) >= six_am_bef:
                        six_am_bef_contacts.append(contact.attributes['antName'].value)
                  elif datetime.strptime(contact.attributes['start'].value, t_format) <= six_am_aft and \
                      datetime.strptime(contact.attributes['end'].value, t_format) >= six_am_aft:
                        six_am_aft_contacts.append(contact.attributes['antName'].value)
                  elif datetime.strptime(contact.attributes['end'].value, t_format) >= burn_times[0] and \
                      datetime.strptime(contact.attributes['start'].value, t_format) < burn_times[-1]:
                        if len(burn_times) > 2:
                              i = 0
                              while i in range(0, len(burn_times)):
                                    if datetime.strptime(contact.attributes['start'].value, t_format) <= burn_times[i] and \
                                        datetime.strptime(contact.attributes['end'].value, t_format) >= burn_times[i]:
                                          burn_contacts.append(contact.attributes['antName'].value)
                                    else:
                                          burn_contacts.append('N/A')
                                    i += 2
                                    
                        else:
                              if datetime.strptime(contact.attributes['start'].value, t_format) <= burn_times[0] and \
                                                                             datetime.strptime(contact.attributes['end'].value, t_format) >= burn_times[0]:
                                          burn_contacts.append(contact.attributes['antName'].value)                         
                  if datetime.strptime(contact.attributes['start'].value, t_format) > burn_times[-1] and \
                       datetime.strptime(contact.attributes['end'].value, t_format) - datetime.strptime(contact.attributes['start'].value, t_format) \
                           > timedelta(minutes = 10):
                        contacts_aft_burn.append(contact.attributes['antName'].value)
                        contacts_aft_burn.append(contact.attributes['start'].value)
                        contacts_aft_burn.append(contact.attributes['end'].value)
      index = 1
      list_contact_start_aft = []
      while index < len(contacts_aft_burn):
            list_contact_start_aft.append(datetime.strptime(contacts_aft_burn[index], t_format))
            index += 3
      cleanup_path_start_time = heapq.nsmallest(2, list_contact_start_aft)[-1]
      pointer = 1
      while pointer < len(contacts_aft_burn):
            if datetime.strptime(contacts_aft_burn[pointer], t_format) == cleanup_path_start_time:
                  cleanup_contacts = contacts_aft_burn[pointer-1:pointer + 2]
            pointer += 3
      return [six_am_bef_contacts, six_am_aft_contacts, burn_contacts, cleanup_contacts]

#######################################################################################################################

def checkThurstPair(thrusters, flightModel):
      #path_CSR =  'C:/Python35/SK_project'
      path_CSR =  'C:/Python35/SK_project'
      list_CSR = []
      for i in os.listdir(path_CSR):
            if os.path.isfile(os.path.join(path_CSR,i)) and 'GB2 CSR' in i:
                  list_CSR.append(i)
            
      list_CSR_date = []
      for j in range(0, len(list_CSR)):
            list_CSR_date.append(datetime.fromtimestamp(os.path.getmtime(path_CSR + '/' + list_CSR[j])))

      index_latest_CSR = list_CSR_date.index(max(list_CSR_date))
      sel_CSR = path_CSR + '/' + list_CSR[index_latest_CSR]

      CSR_file = openpyxl.load_workbook(sel_CSR)
      sheet = CSR_file.get_sheet_by_name('Sheet2')
      for i in range(4, 200):
            if sheet.cell(row=3, column=i).value is not None:
                  if flightModel in sheet.cell(row=3, column=i).value:
                        col = i
      if sheet.cell(row = int(thrusters[0]) + 13, column = col).value is not None and \
         sheet.cell(row = int(thrusters[1]) + 13, column = col).value is not None:
            if 'DEGRADED' in sheet.cell(row = int(thrusters[0]) + 13, column = col).value and \
               'DEGRADED' in sheet.cell(row = int(thrusters[1]) + 13, column = col).value:
                  return False
      else:
                  return True
      
#######################################################################################################            

def checkSunIntrusion(burn_times, sun_intru_entry, sun_intru_exit):
      if len(sun_intru_entry) != 0:
            if len(sun_intru_entry) == len(sun_intru_exit):
                  same_length = 1
            else:
                  same_length = 0
            for j in range(0, len(sun_intru_entry)):
                  if same_length == 1 and sun_intru_entry[0] < sun_intru_exit[0]:
                        for i in range(0, len(burn_times)):
                              if sun_intru_entry[j] < burn_times[i] < sun_intru_exit[j]:
                                    return False
                              elif 2*i in range(0, len(burn_times)):
                                    if burn_times[2*i] < sun_intru_entry[j] < burn_times[2*i + 1]:
                                          return False
                                    elif burn_times[2*i] < sun_intru_exit[j] < burn_times[2*i + 1]:
                                          return False
                              else:
                                    return True
                  elif sun_intru_entry[0] > sun_intru_exit[0]:
                        for i in range(0, len(burn_times)):
                              if sun_intru_entry[j] < burn_times[i] < sun_intru_exit[j+1]:
                                    return False
                              elif burn_times[i] < sun_intru_exit[0]:
                                    return False
                              elif 2*i in range(0, len(burn_times)):
                                    if burn_times[2*i] < sun_intru_entry[j] < burn_times[2*i + 1]:
                                          return False
                                    elif burn_times[2*i] < sun_intru_exit[j] < burn_times[2*i + 1]:
                                          return False
                              else:
                                    return True
                  elif sun_intru_entry[0] < sun_intru_exit[0]:
                        for i in range(0, len(burn_times)):
                              if sun_intru_entry[j] < burn_times[i] < sun_intru_exit[j]:
                                    return False
                              elif burn_times[i] > sun_intru_entry[-1]:
                                    return False
                              elif 2*i in range(0, len(burn_times)):
                                    if burn_times[2*i] < sun_intru_entry[j] < burn_times[2*i + 1]:
                                          return False
                                    elif burn_times[2*i] < sun_intru_exit[j] < burn_times[2*i + 1]:
                                          return False
                              else:
                                    return True
      else:
            return True

#######################################################################################################

def findNoons(burn_times, noon, NBK_flag):
      noon_matters = []
      index = []
      for noon_i in range(0, len(noon)):
            if len(burn_times) > 2:
                  if burn_times[0] < noon[noon_i] and noon[noon_i] < burn_times[-1]:
                        index.append(noon_i)
                        noon_matters.append(noon[noon_i])
            else:
                  if burn_times[0] < noon[noon_i]:
                        index.append(noon_i)
                  
      noon_matters.insert(0, noon[index[0]-1])
      FSS_CL_interval = []
      if NBK_flag == 0:
          for i in range(0, len(noon_matters)):
              FSS_CL_interval.append(noon_matters[i] + timedelta(minutes = 9.5))
              FSS_CL_interval.append(noon_matters[i] + timedelta(minutes = 14.5))
      elif NBK_flag == 1:
          for i in range(0, len(noon_matters)):
              FSS_CL_interval.append(noon_matters[i] + timedelta(minutes = 9.5))
              FSS_CL_interval.append(noon_matters[i] + timedelta(minutes = 19.5))

      return FSS_CL_interval

#########################################################################################################

def findEclipses(burn_times, eclipse_exit, NBK_flag):
      eclipse_matters = []
      index = []
      for eclipse_i in range(0, len(eclipse_exit)):
            if len(burn_times) > 2:
                  if burn_times[0] < eclipse_exit[eclipse_i] and eclipse_exit[eclipse_i] < burn_times[-1]:
                        index.append(eclipse_i)
                        eclipse_matters.append(eclipse_exit[eclipse_i])
            else:
                  if burn_times[0] < eclipse_exit[eclipse_i]:
                        index.append(eclipse_i)
                        
      eclipse_matters.insert(0, eclipse_exit[index[0]-1])
      FSS_eclipse_interval = []
      if NBK_flag == 0:
          for i in range(0, len(eclipse_matters)):
              FSS_eclipse_interval.append(eclipse_matters[i] + timedelta(minutes = 4.5))
              FSS_eclipse_interval.append(eclipse_matters[i] + timedelta(minutes = 9.5))
      elif NBK_flag == 1:
          for i in range(0, len(eclipse_matters)):
              FSS_eclipse_interval.append(eclipse_matters[i] + timedelta(minutes = 4.5))
              FSS_eclipse_interval.append(eclipse_matters[i] + timedelta(minutes = 14.5))
      
      return FSS_eclipse_interval

#########################################################################################################

def checkCL(burn_times, FSS_CL_interval):
      i = 0
      while i in range(0, len(burn_times)):
            if FSS_CL_interval[i] < burn_times[i] < FSS_CL_interval[i+1] or \
               FSS_CL_interval[i] < burn_times[i+1] < FSS_CL_interval[i+1]:
                  i += 2
                  return False
            else:
                  return True

#########################################################################################################

def checkEclipse(burn_times, FSS_eclipse_interval):
      i = 0
      while i in range(0, len(burn_times)):
            if FSS_eclipse_interval[i] < burn_times[i] < FSS_eclipse_interval[i+1] or \
               FSS_eclipse_interval[i] < burn_times[i+1] < FSS_eclipse_interval[i+1]:
                  i += 2
                  return False
            else:
                  return True

#########################################################################################################
                                          
# def quit_program(self):
      
############################################ Main Body ##############################################################

flightModel = input('Enter the flight model number(M***): ')
Current_T = datetime.utcnow()
t_format = '%Y/%m/%d %H:%M:%S.%f'


if flightModel in ['M076', 'M078', 'M092', 'M097']:
    NBK_flag = 1
else:
    NBK_flag = 0

## Looking for latest GB2_MANON file
# path_MANON =  'N:/IOdata/FDS/' + flightModel
path_MANON =  'C:/Python35/SK_project'
list_MANON = []
for i in os.listdir(path_MANON):
      if os.path.isfile(os.path.join(path_MANON,i)) and 'GB2_MANNOM_' in i:
            list_MANON.append(i)

list_MANON_date = []
for j in range(0, len(list_MANON)):
      list_MANON_date.append(datetime.fromtimestamp(os.path.getmtime(path_MANON + '/' + list_MANON[j])))

index_latest_MANON = list_MANON_date.index(max(list_MANON_date))
sel_MANON = path_MANON + '/' + list_MANON[index_latest_MANON] # name of Latest modified MANON file

burn_times = findSKburnTimes(sel_MANON, t_format) # call the function to find SK burn times (list of burn start times and end times)


#

#if Current_T > burn_times[0]: # for testing
if Current_T < burn_times[0]: # the burn start time is in the future
    six_am_bef = []
    six_am_aft = []
    LCZ_interval = []
    if NBK_flag == 0: 
        [sel_orbit, six_am, eclipse_entry, eclipse_exit, noon, sun_intru_entry, sun_intru_exit] = \
               findOrbitEvent(flightModel, burn_times, t_format, NBK_flag)
        
        for i in range(0, len(six_am)):
            if (burn_times[0] - timedelta(hours = 2)) < six_am[i] < burn_times[0]:
                six_am_bef.append(six_am[i])
            elif (burn_times[-1] + timedelta(hours = 3))  < six_am[i] < (burn_times[-1] + timedelta(hours = 5)):
                six_am_aft.append(six_am[i])
                     
        six_am_bef = max(six_am_bef)
        six_am_aft  = max(six_am_aft)               
        [six_am_bef_contacts, six_am_aft_contacts, burn_contacts, cleanup_contacts] = \
            findGroundContact(flightModel, burn_times, six_am_bef, six_am_aft, t_format, NBK_flag)
    else:
        [LCZ_entry, LCZ_exit, eclipse_entry, eclipse_exit, noon, sun_intru_entry, sun_intru_exit] = \
               findOrbitEvent(flightModel, burn_times, t_format, NBK_flag)

        for i in range(0, len(LCZ_entry)):
            if LCZ_exit[i] >= burn_times[0] and LCZ_entry[i] <= burn_times[-1]:
                LCZ_interval.append([LCZ_entry[i], LCZ_exit[i]])
        if len(burn_times) == 2:
            if LCZ_interval:
                tk.Tk().withdraw()
                msg.showinfo('message', \
                   'LCZ zone violation!!!')
        else:
            burn_index = 0
            while burn_index in range(0, len(burn_times)):
                for LCZ_index in range(0, len(LCZ_interval)):
                    if LCZ_interval[LCZ_index][1] < burn_times[burn_index+1] and \
                        LCZ_interval[LCZ_index][2] > burn_times[burn_index]:
                        LCZ_violation = 1
                if LCZ_violation == 1:
                    tk.Tk().withdraw()
                    msg.showinfo('message', \
                                'LCZ zone violation!!!')
                    break        
                burn_index += 2            
        [six_am_bef_contacts, six_am_aft_contacts, burn_contacts, cleanup_contacts] = \
            findGroundContact(flightModel, burn_times, six_am_bef, six_am_aft, t_format, NBK_flag)             
        

    FSS_CL_interval = findNoons(burn_times, noon, NBK_flag)
    FSS_eclipse_interval = findEclipses(burn_times, eclipse_exit, NBK_flag)


                     
    [thrusters, bias_RPY, delta_V]  = findburnInfos(sel_MANON)
    #SK_checklist_file = openpyxl.load_workbook('N:/sateng/Bus Subsystems/Subsys_AOCS/SK Maneuvers/YYYY MM DD MXXX Automated NMSK checklist.xlsx')
    SK_checklist_file = openpyxl.load_workbook('C:/Python35/SK_project/YYYY MM DD MXXX Automated NMSK checklist.xlsx')
      
      
      ## Fill out the times for the burns, cleanup path, and 6am before and after the burn.
      ######################################## One Burn ############################################################  
    if len(burn_times) == 2:
          sheet = SK_checklist_file.get_sheet_by_name('1 Burn')
          sheet['C16'].value = six_am_bef.strftime(t_format)[0:23]
          if len(six_am_bef_contacts) != 0:
                sheet['E16'].value = six_am_bef_contacts[0]
          else:
                sheet['E16'].value = 'N/A'
          sheet['C67'].value = six_am_aft.strftime(t_format)[0:23]
          if len(six_am_aft_contacts) != 0:
                sheet['E67'].value = six_am_aft_contacts[0]
          else:
                sheet['E67'].value = 'N/A'
          sheet['C41'].value = burn_times[0].strftime(t_format)[0:23]
          sheet['C42'].value = burn_times[1].strftime(t_format)[0:23]
          if len(burn_contacts) != 0:
                sheet['E41'].value = burn_contacts[0]
          else:
                sheet['E41'].value = 'N/A'
          sheet['C65'].value = cleanup_contacts[1]
          sheet['C66'].value = cleanup_contacts[1]
          if len(cleanup_contacts) != 0:
                sheet['E65'].value = cleanup_contacts[0]
                sheet['E66'].value = cleanup_contacts[0]
          else:
                sheet['E65'].value = 'N/A'
                sheet['E66'].value = 'N/A'
    
    ######################################## Two Burns ############################################################                  
    elif len(burn_times) == 4:
          sheet = SK_checklist_file.get_sheet_by_name('2 Burn')
          sheet['C16'].value = six_am_bef.strftime(t_format)[0:23]
          if len(six_am_bef_contacts) != 0:
                sheet['E16'].value = six_am_bef_contacts[0]
          else:
                sheet['E16'].value = 'N/A'
          sheet['C74'].value = six_am_aft.strftime(t_format)[0:23]
          if len(six_am_aft_contacts) != 0:
                sheet['E74'].value = six_am_aft_contacts[0]
          else:
                sheet['E67'].value = 'N/A'
          sheet['C46'].value = burn_times[0].strftime(t_format)[0:23]
          sheet['C47'].value = burn_times[1].strftime(t_format)[0:23]
          sheet['C48'].value = burn_times[2].strftime(t_format)[0:23]
          sheet['C49'].value = burn_times[3].strftime(t_format)[0:23]
          sheet['E46'].value = burn_contacts[0]
          sheet['E48'].value = burn_contacts[1]
          
          sheet['C72'].value = cleanup_contacts[1]
          sheet['C73'].value = cleanup_contacts[1]
          if len(cleanup_contacts) != 0:
                sheet['E72'].value = cleanup_contacts[0]
                sheet['E73'].value = cleanup_contacts[0]
          else:
                sheet['E72'].value = 'N/A'
                sheet['E73'].value = 'N/A'
          
    ######################################## Three Burns ############################################################
    elif len(burn_times) == 6:
          sheet = SK_checklist_file.get_sheet_by_name('3 Burn')
          sheet['C16'].value = six_am_bef.strftime(t_format)[0:23]
          if len(six_am_bef_contacts) != 0:
                sheet['E16'].value = six_am_bef_contacts[0]
          else:
                sheet['E16'].value = 'N/A'
          sheet['C76'].value = six_am_aft.strftime(t_format)[0:23]
          if len(six_am_aft_contacts) != 0:
                sheet['E76'].value = six_am_aft_contacts[0]
          else:
                sheet['E67'].value = 'N/A'
          sheet['C46'].value = burn_times[0].strftime(t_format)[0:23]
          sheet['C47'].value = burn_times[1].strftime(t_format)[0:23]
          sheet['C48'].value = burn_times[2].strftime(t_format)[0:23]
          sheet['C49'].value = burn_times[3].strftime(t_format)[0:23]
          sheet['C50'].value = burn_times[4].strftime(t_format)[0:23]
          sheet['C51'].value = burn_times[5].strftime(t_format)[0:23]
          sheet['E46'].value = burn_contacts[0]
          sheet['E48'].value = burn_contacts[1]
          sheet['E50'].value = burn_contacts[2]
         
          sheet['C74'].value = cleanup_contacts[1]
          sheet['C75'].value = cleanup_contacts[1]
          if len(cleanup_contacts) != 0:
                sheet['E74'].value = cleanup_contacts[0]
                sheet['E75'].value = cleanup_contacts[0]
          else:
                sheet['E74'].value = 'N/A'
                sheet['E75'].value = 'N/A'
    
    ######################################## Four Burns ############################################################                  
    elif len(burn_times) == 8:
          sheet = SK_checklist_file.get_sheet_by_name('4 Burn')
          sheet['C16'].value = six_am_bef.strftime(t_format)[0:23]
          if len(six_am_bef_contacts) != 0:
                sheet['E16'].value = six_am_bef_contacts[0]
          else:
                sheet['E16'].value = 'N/A'
          sheet['C77'].value = six_am_aft.strftime(t_format)[0:23]
          if len(six_am_aft_contacts) != 0:
                sheet['E77'].value = six_am_aft_contacts[0]
          else:
                sheet['E67'].value = 'N/A'
          sheet['C45'].value = burn_times[0].strftime(t_format)[0:23]
          sheet['C46'].value = burn_times[1].strftime(t_format)[0:23]
          sheet['C47'].value = burn_times[2].strftime(t_format)[0:23]
          sheet['C48'].value = burn_times[3].strftime(t_format)[0:23]
          sheet['C49'].value = burn_times[4].strftime(t_format)[0:23]
          sheet['C50'].value = burn_times[5].strftime(t_format)[0:23]
          sheet['C51'].value = burn_times[6].strftime(t_format)[0:23]
          sheet['C52'].value = burn_times[7].strftime(t_format)[0:23]
          sheet['E45'].value = burn_contacts[0]
          sheet['E47'].value = burn_contacts[1]
          sheet['E49'].value = burn_contacts[2]
          sheet['E51'].value = burn_contacts[3]
          
          sheet['C75'].value = cleanup_contacts[1]
          sheet['C76'].value = cleanup_contacts[1]
          if len(cleanup_contacts) != 0:
                sheet['E75'].value = cleanup_contacts[0]
                sheet['E76'].value = cleanup_contacts[0]
          else:
                sheet['E75'].value = 'N/A'
                sheet['E76'].value = 'N/A'
          
    ## Check the pre-conditions for the burns
    sheet['A1'].value = sheet['A1'].value.replace(sheet['A1'].value[0:4], flightModel)
    sheet['B3'].value = sel_MANON
    sheet['E11'].value = 'X'
    if checkThurstPair(thrusters, flightModel) is True:
          sheet['E4'].value = 'X'
    elif checkThurstPair(thrusters, flightModel) is False:
        tk.Tk().withdraw()
        msg.showinfo('message', \
                   'The slected thrusters are degraded!!')
    if float(delta_V) < 0.04:
          sheet['E5'].value = 'X'
    elif float(delta_V) >= 0.04:
        tk.Tk().withdraw()
        msg.showinfo('message', \
                   'Delta V is over the limit!!')
    if checkCL(burn_times, FSS_CL_interval) is True and checkEclipse(burn_times, FSS_eclipse_interval) is True:
          sheet['E8'].value = 'X'
    elif checkCL(burn_times, FSS_CL_interval) is False or checkEclipse(burn_times, FSS_eclipse_interval) is False:
        tk.Tk().withdraw()
        msg.showinfo('message', \
                   'The burn is within 5 minutes after FSS valid!!')      
    if checkSunIntrusion(burn_times, sun_intru_entry, sun_intru_exit) is True:
          sheet['E10'].value = 'X'
    elif checkSunIntrusion(burn_times, sun_intru_entry, sun_intru_exit) is False:
        tk.Tk().withdraw()
        msg.showinfo('message', \
                   'There is a sun intrusion involved!!') 
    if abs(float(bias_RPY[2])) == 180.0:
          sheet['E13'].value = 'Retrograde'
    elif abs(float(bias_RPY[2])) == 0.0:
          sheet['E13'].value = 'Posigrade'
               
    filename = str(burn_times[0].year) + ' ' + str(burn_times[0].month) + ' ' + str(burn_times[0].day) + ' ' + \
               flightModel + ' Automated NMSK checklist.xlsx'
    #SK_checklist_file.save('N:/sateng/Bus Subsystems/Subsys_AOCS/SK Maneuvers/' + flightModel + '/' + filename)
    SK_checklist_file.save('C:/Python35/SK_project/' + flightModel + '/' + filename)
    
else:
      tk.Tk().withdraw()
      msg.showinfo('message', \
                   'The proper MANON file doesn\'t exit.')

      
      
      
